"""
Export service for generating CSV and Excel files from CRL data.

This service handles exporting CRL data in various formats with optional
inclusion of summaries and filtering.
"""

import csv
import io
from typing import List, Optional
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from app.utils.logging_config import get_logger

logger = get_logger(__name__)


class ExportService:
    """Service for exporting CRL data to various formats."""

    # Define the export columns
    EXPORT_COLUMNS = [
        ("ID", "id"),
        ("Application Number", "application_number"),
        ("Company Name", "company_name"),
        ("Letter Date", "letter_date"),
        ("Letter Year", "letter_year"),
        ("Application Type", "application_type"),
        ("Letter Type", "letter_type"),
        ("Approval Status", "approval_status"),
        ("Therapeutic Category", "therapeutic_category"),
        ("Product Name", "product_name"),
        ("Indications", "indications"),
        ("Deficiency Reason", "deficiency_reason"),
        ("Approver Center", "approver_center"),
        ("Approver Name", "approver_name"),
    ]

    SUMMARY_COLUMN = ("Executive summary", "summary")

    @staticmethod
    def _format_value(value) -> str:
        """Format a value for export."""
        if value is None:
            return ""
        elif isinstance(value, list):
            return ", ".join(str(v) for v in value)
        elif isinstance(value, datetime):
            return value.isoformat()
        else:
            return str(value)

    @staticmethod
    def export_to_csv(
        crls: List[dict],
        include_summary: bool = False
    ) -> io.StringIO:
        """
        Export CRLs to CSV format.

        Args:
            crls: List of CRL dictionaries
            include_summary: Whether to include AI summary column

        Returns:
            StringIO object containing CSV data
        """
        logger.info(f"Exporting {len(crls)} CRLs to CSV (include_summary={include_summary})")

        output = io.StringIO()

        # Determine columns
        columns = list(ExportService.EXPORT_COLUMNS)
        if include_summary:
            columns.append(ExportService.SUMMARY_COLUMN)

        # Write CSV
        writer = csv.writer(output)

        # Write header
        writer.writerow([col[0] for col in columns])

        # Write data rows
        for crl in crls:
            row = [ExportService._format_value(crl.get(col[1])) for col in columns]
            writer.writerow(row)

        logger.info(f"✓ Successfully generated CSV with {len(crls)} rows")

        output.seek(0)
        return output

    @staticmethod
    def export_to_excel(
        crls: List[dict],
        include_summary: bool = False
    ) -> io.BytesIO:
        """
        Export CRLs to Excel format.

        Args:
            crls: List of CRL dictionaries
            include_summary: Whether to include AI summary column

        Returns:
            BytesIO object containing Excel data

        Raises:
            ImportError: If openpyxl is not installed
        """
        if not HAS_OPENPYXL:
            raise ImportError(
                "openpyxl is required for Excel export. "
                "Install with: pip install openpyxl"
            )

        logger.info(f"Exporting {len(crls)} CRLs to Excel (include_summary={include_summary})")

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "CRL Export"

        # Determine columns
        columns = list(ExportService.EXPORT_COLUMNS)
        if include_summary:
            columns.append(ExportService.SUMMARY_COLUMN)

        # Style for header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        # Write header
        for col_idx, (header, _) in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font

        # Write data rows
        for row_idx, crl in enumerate(crls, start=2):
            for col_idx, (_, field) in enumerate(columns, start=1):
                value = ExportService._format_value(crl.get(field))
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-size columns (approximate)
        for col_idx, (header, _) in enumerate(columns, start=1):
            # Set column width based on header length (minimum 12, maximum 50)
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = min(max(len(header) + 2, 12), 50)

        # Freeze header row
        ws.freeze_panes = "A2"

        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        logger.info(f"✓ Successfully generated Excel file with {len(crls)} rows")

        return output
